"""
Report Generator — Phase 3/4
Generates all 5 required output artifacts per Блок 3 spec:
  1. executive_summary   — PDF (with AI narrative)
  2. gap_register        — XLSX
  3. risk_register       — XLSX
  4. roadmap             — XLSX (Remediation Roadmap)
  5. evidence_checklist  — XLSX

All data sourced exclusively from engine outputs.
AI used only for narrative sections of Executive Summary.
Human review required before publish (enforced by status workflow).
"""
import io
import json
import uuid
from datetime import datetime, timezone
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
)
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func

from app.models.models import (
    Assessment, Tenant, ControlResult, Gap, Risk,
    RemediationAction, EvidenceLink, Control, Question, EvidenceFile
)
from app.core.config import settings
from app.services.evidence_aggregator import recompute_control_aggregates
from app.services.report_context_builder import build_full_report_context


# ── Colour palette ─────────────────────────────────────────────────────────────
SEVERITY_COLORS = {
    "Critical": "C0392B",
    "High":     "E67E22",
    "Medium":   "F1C40F",
    "Low":      "2ECC71",
}
HEADER_FILL = "1A3A5C"   # dark navy
SUBHEADER_FILL = "2980B9"  # blue
ALT_ROW_FILL = "EBF5FB"


def _severity_fill(severity: str) -> PatternFill:
    hex_color = SEVERITY_COLORS.get(severity, "AAAAAA")
    return PatternFill("solid", fgColor=hex_color)


def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=HEADER_FILL)


def _alt_fill() -> PatternFill:
    return PatternFill("solid", fgColor=ALT_ROW_FILL)


def _thin_border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_header_row(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _header_fill()
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()


def _auto_width(ws, min_width=12, max_width=60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_width, max(min_width, max_len + 2))


# ── XLSX: Gap Register ─────────────────────────────────────────────────────────

async def generate_gap_register(
    assessment: Assessment,
    tenant: Tenant,
    db: AsyncSession,
) -> bytes:
    gaps_result = await db.execute(
        select(Gap, Control)
        .join(Control, Control.id == Gap.control_id)
        .where(Gap.assessment_id == assessment.id)
        .order_by(Gap.severity.desc(), Control.control_code)
    )
    rows = gaps_result.all()

    # Evidence per control
    ev_result = await db.execute(
        select(EvidenceLink.control_id)
        .where(EvidenceLink.assessment_id == assessment.id, EvidenceLink.control_id.is_not(None))
    )
    controls_with_evidence = {r[0] for r in ev_result.all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Gap Register"

    headers = [
        "Control ID", "HIPAA Reference", "Control Title", "Category",
        "Status", "Severity", "Gap Description", "Recommended Remediation",
        "Evidence Provided"
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 30

    for i, (gap, control) in enumerate(rows, start=2):
        evidence_provided = "Yes" if control.id in controls_with_evidence else "No"
        row = [
            control.control_code,
            f"45 CFR 164 — {control.category}",
            control.title,
            control.category,
            gap.status_source,
            gap.severity,
            gap.description,
            gap.recommended_remediation or "",
            evidence_provided,
        ]
        ws.append(row)

        # Alternate row fill
        fill = _alt_fill() if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=col)
            cell.fill = fill
            cell.border = _thin_border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Severity cell color
        sev_cell = ws.cell(row=i, column=6)
        sev_cell.fill = _severity_fill(gap.severity)
        sev_cell.font = Font(bold=True, color="FFFFFF")

    _auto_width(ws)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── XLSX: Risk Register ────────────────────────────────────────────────────────

async def generate_risk_register(
    assessment: Assessment,
    tenant: Tenant,
    db: AsyncSession,
) -> bytes:
    risks_result = await db.execute(
        select(Risk, Gap, Control)
        .join(Gap, Gap.id == Risk.gap_id)
        .join(Control, Control.id == Gap.control_id)
        .where(Risk.assessment_id == assessment.id)
        .order_by(Risk.severity.desc(), Control.control_code)
    )
    rows = risks_result.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Risk Register"

    headers = [
        "Risk ID", "Related Control", "Control Title", "Severity",
        "Risk Description", "Rationale", "Recommended Action"
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 30

    for i, (risk, gap, control) in enumerate(rows, start=2):
        row = [
            f"RSK-{i - 1:03d}",
            control.control_code,
            control.title,
            risk.severity,
            risk.description,
            risk.rationale or "",
            gap.recommended_remediation or "",
        ]
        ws.append(row)

        fill = _alt_fill() if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=col)
            cell.fill = fill
            cell.border = _thin_border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        sev_cell = ws.cell(row=i, column=4)
        sev_cell.fill = _severity_fill(risk.severity)
        sev_cell.font = Font(bold=True, color="FFFFFF")

    _auto_width(ws)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── XLSX: Remediation Roadmap ──────────────────────────────────────────────────

async def generate_roadmap(
    assessment: Assessment,
    tenant: Tenant,
    db: AsyncSession,
) -> bytes:
    rem_result = await db.execute(
        select(RemediationAction, Gap, Control)
        .join(Gap, Gap.id == RemediationAction.gap_id)
        .join(Control, Control.id == Gap.control_id)
        .where(RemediationAction.assessment_id == assessment.id)
        .order_by(RemediationAction.priority, Control.control_code)
    )
    rows = rem_result.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Remediation Roadmap"

    # Priority order annotation
    PRIORITY_PHASE = {
        "Critical": "30-Day Sprint",
        "High":     "60-Day Window",
        "Medium":   "90-Day Window",
        "Low":      "Ongoing / Backlog",
    }

    headers = [
        "Action ID", "Related Control", "Control Title", "Priority",
        "Phase", "Effort", "Type", "Description", "Dependency", "Template Ref"
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 30

    for i, (rem, gap, control) in enumerate(rows, start=2):
        row = [
            f"ACT-{i - 1:03d}",
            control.control_code,
            control.title,
            rem.priority,
            PRIORITY_PHASE.get(rem.priority, ""),
            rem.effort,
            rem.remediation_type,
            rem.description,
            rem.dependency or "",
            rem.template_reference or "",
        ]
        ws.append(row)

        fill = _alt_fill() if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=col)
            cell.fill = fill
            cell.border = _thin_border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        pri_cell = ws.cell(row=i, column=4)
        pri_cell.fill = _severity_fill(rem.priority)
        pri_cell.font = Font(bold=True, color="FFFFFF")

    _auto_width(ws)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── XLSX: Evidence Checklist ───────────────────────────────────────────────────

async def generate_evidence_checklist(
    assessment: Assessment,
    tenant: Tenant,
    db: AsyncSession,
) -> bytes:
    # All controls
    ctrl_result = await db.execute(
        select(Control, ControlResult)
        .join(ControlResult, (
            (ControlResult.control_id == Control.id) &
            (ControlResult.assessment_id == assessment.id)
        ), isouter=True)
        .where(Control.controlset_version_id == assessment.controlset_version_id)
        .order_by(Control.control_code)
    )
    rows = ctrl_result.all()

    # Evidence per control
    ev_result = await db.execute(
        select(EvidenceLink.control_id, EvidenceFile.file_name)
        .join(EvidenceFile, EvidenceFile.id == EvidenceLink.evidence_file_id)
        .where(EvidenceLink.assessment_id == assessment.id, EvidenceLink.control_id.is_not(None))
    )
    ev_map: dict[str, list[str]] = {}
    for ctrl_id, fname in ev_result.all():
        ev_map.setdefault(ctrl_id, []).append(fname)

    wb = Workbook()
    ws = wb.active
    ws.title = "Evidence Checklist"

    EXPECTED_EVIDENCE = {
        "Policy": "Policy document (PDF/DOCX)",
        "Technical": "Screenshot / configuration export / vendor certificate",
        "Process": "Process document / checklist / training record",
    }

    headers = [
        "Control ID", "Control Title", "Category", "Severity", "Status",
        "Expected Evidence", "Evidence Provided", "Files Uploaded", "Notes", "Status Impact"
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 30

    for i, (control, cr) in enumerate(rows, start=2):
        ctrl_status = cr.status if cr else "Unknown"
        files = ev_map.get(control.id, [])
        ev_provided = "Yes" if files else "No"
        files_str = "; ".join(files) if files else "—"

        # Expected evidence hint
        from app.services.engine import CONTROL_TEMPLATES, REMEDIATION_TEMPLATES
        tmpl_id = CONTROL_TEMPLATES.get(control.control_code, "")
        tmpl = REMEDIATION_TEMPLATES.get(tmpl_id, {})
        ev_type = tmpl.get("type", "Process")
        expected = EXPECTED_EVIDENCE.get(ev_type, "Documentation")

        # Status impact
        if ctrl_status == "Fail":
            impact = "High — gap identified"
        elif ctrl_status == "Partial":
            impact = "Medium — partial compliance"
        elif ctrl_status == "Unknown":
            impact = "Medium — status unverified"
        else:
            impact = "None — control passing"

        row = [
            control.control_code, control.title, control.category, control.severity,
            ctrl_status, expected, ev_provided, files_str, "", impact
        ]
        ws.append(row)

        fill = _alt_fill() if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=col)
            cell.fill = fill
            cell.border = _thin_border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        if ctrl_status not in ("Pass",):
            ws.cell(row=i, column=5).fill = _severity_fill(control.severity)
            ws.cell(row=i, column=5).font = Font(bold=True, color="FFFFFF")

    _auto_width(ws)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── AI Narrative ───────────────────────────────────────────────────────────────

def _build_legacy_prompt(
    tenant: Tenant,
    assessment: Assessment,
    stats: dict,
    top_gaps: list,
    ai_tone: str,
) -> str:
    """Legacy prompt (stats + top_gaps only). Used when full_context is not available."""
    tone_instruction = {
        "neutral": "Use a professional, neutral tone.",
        "executive": "Use an executive-friendly tone. Focus on business risk and strategic priorities.",
        "plain_language": "Use simple, plain language. Avoid technical jargon. Assume non-technical readers.",
    }.get(ai_tone, "Use a professional, neutral tone.")

    critical_gaps = [g for g in top_gaps if g.severity == "Critical"]
    high_gaps = [g for g in top_gaps if g.severity == "High"]

    return f"""You are a HIPAA compliance expert writing an Executive Compliance Summary.

Organization: {tenant.name}
Assessment date: {assessment.submitted_at.strftime('%B %d, %Y') if assessment.submitted_at else 'N/A'}

Engine results:
- Total controls assessed: {stats.get('total', 0)}
- Pass: {stats.get('pass', 0)}
- Fail: {stats.get('fail', 0)}
- Partial: {stats.get('partial', 0)}
- Unknown: {stats.get('unknown', 0)}
- Total gaps identified: {stats.get('gaps', 0)}
- Critical gaps: {len(critical_gaps)}
- High severity gaps: {len(high_gaps)}

Top gaps:
{chr(10).join(f'- [{getattr(g, "severity", "")}] {(g.description or "")[:120]}' for g in top_gaps[:8])}

{tone_instruction}

Write a concise executive narrative (4-6 paragraphs) covering:
1. Overall security posture assessment
2. Key findings summary
3. Priority risk areas requiring immediate attention
4. Recommended next steps (30-60-90 day)

Do NOT invent findings. Base everything strictly on the data provided.
Do NOT use bullet points in the narrative — use flowing prose only.
Keep it under 600 words."""


def _build_full_analysis_prompt(context: dict, ai_tone: str) -> str:
    """
    Полный промпт с тремя источниками данных.
    Использует реальные поля из ControlResult (engine_status, engine_rationale)
    и ControlEvidenceAggregate (evidence_status, evidence_avg_strength, evidence_findings).
    """
    tenant = context["tenant"]
    assessment = context["assessment"]
    controls = context["controls"]
    gaps = context["top_gaps"]
    ev_summary = context["evidence_summary"]
    prev = context.get("previous_assessment")

    tone_map = {
        "neutral": "Professional and objective tone.",
        "formal": "Formal regulatory language appropriate for external auditors.",
        "supportive": "Constructive and encouraging tone, focused on actionable improvement.",
    }
    tone_instruction = tone_map.get(ai_tone, tone_map["neutral"])

    gap_controls = [c for c in controls if c["final_status"] == "gap"]
    partial_controls = [c for c in controls if c["final_status"] == "partial"]

    gap_lines = []
    for c in gap_controls[:12]:
        line = f"  [{c['engine_severity']}] {c['control_id']} — {c['control_name']}"
        if c.get("engine_rationale"):
            line += f"\n    Engine: {c['engine_rationale'][:120]}"
        if c.get("evidence_findings"):
            line += f"\n    Evidence findings: {'; '.join(str(f) for f in c['evidence_findings'][:2])}"
        gap_lines.append(line)

    gap_section = "\n".join(gap_lines) if gap_lines else "  None identified"

    partial_lines = [
        f"  {c['control_id']} — {c['control_name']}: "
        f"evidence_strength={c.get('evidence_avg_strength') or 'n/a'}, "
        f"engine={c['engine_status']}"
        for c in partial_controls[:8]
    ]
    partial_section = "\n".join(partial_lines) if partial_lines else "  None"

    delta_section = ""
    if prev:
        delta = assessment["score_percent"] - prev["score_percent"]
        sign = "+" if delta >= 0 else ""
        delta_section = f"""
REMEDIATION PROGRESS (compared to previous assessment on {prev['assessment_date']}):
  Previous score: {prev['score_percent']}%  →  Current score: {assessment['score_percent']}% ({sign}{delta}%)
  Previous gap count: {prev['total_gaps']}  →  Current gap count: {assessment['gaps']}
  Previously identified gaps (sample): {', '.join(prev.get('gap_descriptions', [])[:4])}
"""

    agent_section = ""
    agent = context.get("agent_snapshot")
    if agent:
        received = agent.get("received_at_utc") or "unknown"
        rid = agent.get("receipt_id") or ""
        version = agent.get("agent_version") or ""
        manifest = agent.get("manifest_payload") or {}
        snapshot = agent.get("snapshot_data") or {}
        manifest_str = json.dumps(manifest, ensure_ascii=False)[:1500] if manifest else "—"
        snapshot_str = json.dumps(snapshot, ensure_ascii=False)[:2000] if snapshot else "—"
        agent_section = f"""
DATA FROM LOCAL AGENT (included in this report — do not ignore):
  Last received: {received}
  Receipt ID: {rid}
  Agent version: {version}
  Manifest (package metadata): {manifest_str}
  Snapshot (anonymized agent-reported data): {snapshot_str}
Use this block in your narrative: summarize what the agent reported and how it supports or contrasts with questionnaire and evidence. If no agent data is present above, skip the Agent-Reported Data section.
"""

    return f"""You are a senior HIPAA Security Rule compliance analyst preparing an Executive Compliance Summary.

{tone_instruction}

═══════════════════════════════════════════════
ORGANIZATION: {tenant['name']}
Assessment Date: {assessment.get('submitted_at') or 'Not specified'}
═══════════════════════════════════════════════

OVERALL COMPLIANCE SCORE: {assessment['score_percent']}%
  Controls assessed:  {assessment['total_controls']}
  Passed:             {assessment['passed']}
  Partial compliance: {assessment['partial']}
  Gaps identified:    {assessment['gaps']}

EVIDENCE QUALITY SUMMARY (from AI document analysis):
  Controls with uploaded evidence:  {ev_summary['controls_with_evidence']}
  Strong / Adequate evidence:       {ev_summary['strong_adequate']}
  Weak evidence:                    {ev_summary['weak']}
  Insufficient / Mismatch:          {ev_summary['insufficient']}
  Not analyzed (no upload):         {ev_summary['not_analyzed']}
{delta_section}
CONTROLS WITH GAPS ({len(gap_controls)} total):
{gap_section}

CONTROLS WITH PARTIAL COMPLIANCE ({len(partial_controls)} total):
{partial_section}

TOP PRIORITY GAPS BY SEVERITY:
{chr(10).join(f"  [{g['severity'].upper()}] {g['description']}" + (f" → {g.get('recommended_remediation', '')}" if g.get('recommended_remediation') else '') for g in gaps[:8])}
{agent_section}
═══════════════════════════════════════════════
WRITE AN EXECUTIVE COMPLIANCE SUMMARY with these sections:

**Executive Summary** (3–4 sentences)
Overall compliance posture, key risk level, most critical finding.

**Compliance Score Analysis**
Interpret {assessment['score_percent']}% in HIPAA regulatory context.
What is the organization's exposure at this score level?

**Evidence Quality Assessment**
Comment on the strength of documentation provided.
{ev_summary['strong_adequate']} controls have strong/adequate evidence.
{ev_summary['weak']} controls have weak evidence that needs strengthening.
{ev_summary['not_analyzed']} controls have no evidence uploaded.
What does this mean for audit readiness?
{"**Agent-Reported Data** — Summarize what the local clinic agent reported (see DATA FROM LOCAL AGENT above). How does it support or contrast with the questionnaire and evidence? If no agent data was provided, omit this section." if agent else ""}

**Critical Findings**
Top 3–5 gaps with HIPAA regulatory citations (45 CFR 164.xxx).
For each: what is missing, what is the compliance risk.

**Remediation Priorities**
  Immediate (0–30 days): highest severity items
  Short-term (30–90 days): high severity items
  Ongoing: medium/low severity items
{"**Progress Since Last Assessment**" + chr(10) + "Comment on improvements and remaining gaps." if prev else ""}

Format: professional narrative paragraphs (no bullet points in main sections).
Use the section headers exactly as shown above.
Maximum 750 words total.
═══════════════════════════════════════════════
"""


async def generate_ai_narrative(
    tenant: Tenant,
    assessment: Assessment,
    stats: dict,
    top_gaps: list,
    ai_tone: str = "neutral",
    full_context: dict | None = None,
) -> str:
    """
    Calls Anthropic Claude API for executive summary narrative.
    Если full_context передан — Claude получает полный контекст.
    Если нет — fallback на legacy промпт (stats + top_gaps).
    Falls back to template if LLM not configured or on error.
    """
    if not settings.LLM_ENABLED or not settings.ANTHROPIC_API_KEY:
        return _template_narrative(tenant, assessment, stats, top_gaps)

    try:
        import anthropic
        import logging

        client = anthropic.Anthropic(api_key=settings.ANTHROPIC_API_KEY)

        if full_context:
            prompt = _build_full_analysis_prompt(full_context, ai_tone)
            max_tokens = 2000
        else:
            prompt = _build_legacy_prompt(tenant, assessment, stats, top_gaps, ai_tone)
            max_tokens = 900

        message = client.messages.create(
            model=settings.LLM_MODEL,
            max_tokens=max_tokens,
            messages=[{"role": "user", "content": prompt}],
        )
        return message.content[0].text

    except Exception as e:
        logging.getLogger(__name__).error(
            "Claude API error in generate_ai_narrative: %s", e
        )
        return _template_narrative(tenant, assessment, stats, top_gaps) + (
            f"\n\n[AI narrative unavailable: {e}]"
        )


def _template_narrative(tenant, assessment, stats, top_gaps) -> str:
    total = stats.get("total", 0)
    passing = stats.get("pass", 0)
    failing = stats.get("fail", 0)
    partial = stats.get("partial", 0)
    unknown = stats.get("unknown", 0)
    gaps = stats.get("gaps", 0)

    pass_pct = round(passing / total * 100) if total else 0
    critical_count = sum(1 for g in top_gaps if g.severity == "Critical")
    high_count = sum(1 for g in top_gaps if g.severity == "High")

    return f"""This HIPAA Security Rule Readiness Assessment was conducted for {tenant.name} \
to evaluate compliance with the requirements of 45 CFR Part 164 Subparts C (Security Rule). \
The assessment covers Administrative, Physical, and Technical Safeguards as well as \
Vendor and Third-Party Management obligations. This report reflects a point-in-time readiness \
evaluation and does not constitute a formal HIPAA audit or legal opinion.

Of the {total} controls assessed, {passing} ({pass_pct}%) were found to be in place. \
{failing} controls were identified as non-compliant (Fail), {partial} as partially compliant, \
and {unknown} could not be verified due to insufficient information. A total of {gaps} \
compliance gaps were identified across all safeguard categories.

{"The assessment identified " + str(critical_count) + " Critical-severity findings requiring immediate attention. " if critical_count > 0 else ""}{"Additionally, " + str(high_count) + " High-severity findings present significant compliance risk. " if high_count > 0 else ""}These findings represent areas where electronic Protected Health Information (ePHI) may be at risk of unauthorized access, disclosure, or loss.

Priority remediation actions have been organized into a 30-60-90 day roadmap included in the \
accompanying Remediation Roadmap document. Critical and High severity gaps should be addressed \
within the 30-day sprint to reduce immediate risk exposure. Medium severity findings should be \
scheduled within 60 days, and remaining items incorporated into ongoing compliance operations.

Management is advised to review the Gap Register and Risk Register documents in detail, \
assign remediation owners, and establish a tracking mechanism for closure. A follow-up \
assessment is recommended within 90 days to measure remediation progress.

DISCLAIMER: This assessment is based on self-reported information provided by the organization. \
Findings are dependent on the accuracy and completeness of responses. This report does not \
guarantee HIPAA compliance and should not be used as a substitute for a formal compliance audit \
by a qualified professional."""


# ── PDF: Executive Summary ─────────────────────────────────────────────────────

async def generate_executive_summary(
    assessment: Assessment,
    tenant: Tenant,
    db: AsyncSession,
    include_ai: bool = True,
    ai_tone: str = "neutral",
) -> bytes:
    # Gather stats from DB
    cr_result = await db.execute(
        select(ControlResult).where(ControlResult.assessment_id == assessment.id)
    )
    control_results = cr_result.scalars().all()

    stats = {
        "total": len(control_results),
        "pass": sum(1 for r in control_results if r.status == "Pass"),
        "fail": sum(1 for r in control_results if r.status == "Fail"),
        "partial": sum(1 for r in control_results if r.status == "Partial"),
        "unknown": sum(1 for r in control_results if r.status == "Unknown"),
    }

    gaps_result = await db.execute(
        select(Gap).where(Gap.assessment_id == assessment.id)
        .order_by(Gap.severity.desc())
        .limit(15)
    )
    top_gaps = gaps_result.scalars().all()
    stats["gaps"] = len(top_gaps)

    if include_ai and getattr(settings, "LLM_ENABLED", False):
        try:
            await recompute_control_aggregates(
                assessment_id=str(assessment.id),
                tenant_id=str(assessment.tenant_id),
                db=db,
            )
        except Exception:
            pass

    full_context = None
    if include_ai and getattr(settings, "LLM_ENABLED", False):
        try:
            full_context = await build_full_report_context(
                assessment_id=str(assessment.id),
                tenant_id=str(assessment.tenant_id),
                db=db,
            )
        except Exception:
            full_context = None

    narrative = await generate_ai_narrative(
        tenant, assessment, stats, top_gaps, ai_tone,
        full_context=full_context,
    )

    # Build PDF
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=letter,
        rightMargin=inch,
        leftMargin=inch,
        topMargin=inch,
        bottomMargin=inch,
    )

    styles = getSampleStyleSheet()
    story = []

    # Title block
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"],
        fontSize=18, spaceAfter=6, textColor=colors.HexColor("#1A3A5C")
    )
    subtitle_style = ParagraphStyle(
        "Subtitle", parent=styles["Normal"],
        fontSize=11, spaceAfter=4, textColor=colors.HexColor("#555555")
    )
    heading_style = ParagraphStyle(
        "Heading2", parent=styles["Heading2"],
        fontSize=13, spaceBefore=16, spaceAfter=8,
        textColor=colors.HexColor("#1A3A5C")
    )
    body_style = ParagraphStyle(
        "Body", parent=styles["Normal"],
        fontSize=10, leading=15, spaceAfter=10
    )
    disclaimer_style = ParagraphStyle(
        "Disclaimer", parent=styles["Normal"],
        fontSize=8, textColor=colors.HexColor("#888888"), spaceAfter=4
    )

    story.append(Paragraph("HIPAA Security Rule Readiness Assessment", title_style))
    story.append(Paragraph("Executive Compliance Summary", subtitle_style))
    story.append(Paragraph(f"Organization: <b>{tenant.name}</b>", subtitle_style))
    report_date = assessment.submitted_at.strftime("%B %d, %Y") if assessment.submitted_at else "N/A"
    story.append(Paragraph(f"Assessment Date: {report_date}", subtitle_style))
    story.append(Paragraph(f"Report Generated: {datetime.now().strftime('%B %d, %Y')}", subtitle_style))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#1A3A5C")))
    story.append(Spacer(1, 12))

    # Scope
    scope_meta = assessment.metadata_ or {}
    story.append(Paragraph("Scope & Methodology", heading_style))
    scope_note = scope_meta.get("scope_note", "Full organizational scope — all ePHI systems and workflows.")
    story.append(Paragraph(
        f"This assessment covers HIPAA Security Rule safeguards (Administrative, Physical, Technical) "
        f"and Vendor/Third-Party Management. Scope: {scope_note}. "
        f"Framework: HIPAA Security Rule (45 CFR 164.308 / 310 / 312). "
        f"This is a readiness evaluation, not a formal audit.", body_style
    ))

    # Posture table
    story.append(Paragraph("Overall Security Posture", heading_style))

    posture_data = [
        ["Metric", "Count", "Percentage"],
        ["Controls Assessed", str(stats["total"]), "100%"],
        ["Passing", str(stats["pass"]), f"{round(stats['pass'] / max(stats['total'], 1) * 100)}%"],
        ["Failing", str(stats["fail"]), f"{round(stats['fail'] / max(stats['total'], 1) * 100)}%"],
        ["Partial Compliance", str(stats["partial"]), f"{round(stats['partial'] / max(stats['total'], 1) * 100)}%"],
        ["Status Unknown", str(stats["unknown"]), f"{round(stats['unknown'] / max(stats['total'], 1) * 100)}%"],
        ["Total Gaps Identified", str(stats["gaps"]), "—"],
    ]
    posture_table = Table(posture_data, colWidths=[3 * inch, 1.5 * inch, 1.5 * inch])
    posture_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A3A5C")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#EBF5FB"), colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(posture_table)
    story.append(Spacer(1, 16))

    # Narrative (escape XML entities for ReportLab Paragraph)
    def _escape_para(text: str) -> str:
        if not text:
            return ""
        return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    story.append(Paragraph("Executive Summary", heading_style))
    for para in narrative.split("\n\n"):
        para = _escape_para(para.strip())
        if para:
            story.append(Paragraph(para, body_style))

    # Top findings
    critical_high = [g for g in top_gaps if g.severity in ("Critical", "High")][:8]
    if critical_high:
        story.append(Paragraph("Critical & High Priority Findings", heading_style))
        findings_data = [["Severity", "Control", "Gap Description"]]
        for gap in critical_high:
            # Get control code via query would require async — use description prefix
            d = gap.description or ""
            desc = (d[:100] + "...") if len(d) > 100 else d
            findings_data.append([gap.severity, "—", desc])

        findings_table = Table(findings_data, colWidths=[1 * inch, 1.2 * inch, 4.3 * inch])
        sev_colors_map = {
            "Critical": colors.HexColor("#C0392B"),
            "High": colors.HexColor("#E67E22"),
        }
        findings_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A3A5C")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#EBF5FB"), colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(findings_table)
        story.append(Spacer(1, 16))

    # Disclaimer
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#CCCCCC")))
    story.append(Spacer(1, 8))
    story.append(Paragraph(
        "DISCLAIMER: This report is based on self-reported information. It reflects a point-in-time "
        "readiness evaluation and does not constitute a formal HIPAA audit, legal opinion, or guarantee "
        "of compliance. Engage a qualified HIPAA compliance professional for formal audit purposes.",
        disclaimer_style
    ))
    story.append(Paragraph(
        f"Generated by Summit Range Consulting HIPAA Readiness Platform | {datetime.now().strftime('%Y-%m-%d %H:%M UTC')}",
        disclaimer_style
    ))

    doc.build(story)
    return buf.getvalue()


# ── Main orchestrator ──────────────────────────────────────────────────────────

async def generate_all_reports(
    assessment: Assessment,
    tenant: Tenant,
    db: AsyncSession,
    include_ai: bool = True,
    ai_tone: str = "neutral",
) -> dict[str, bytes]:
    """
    Generate all 5 required report artifacts.
    Returns dict: {file_type: bytes}
    """
    cr_count_result = await db.execute(
        select(func.count()).select_from(ControlResult).where(
            ControlResult.assessment_id == assessment.id
        )
    )
    if cr_count_result.scalar_one() == 0:
        raise ValueError(
            "No control results found for this assessment. Run the compliance engine before generating reports."
        )

    results = {}

    results["executive_summary"] = await generate_executive_summary(
        assessment, tenant, db, include_ai, ai_tone
    )
    results["gap_register"] = await generate_gap_register(assessment, tenant, db)
    results["risk_register"] = await generate_risk_register(assessment, tenant, db)
    results["roadmap"] = await generate_roadmap(assessment, tenant, db)
    results["evidence_checklist"] = await generate_evidence_checklist(assessment, tenant, db)

    return results
